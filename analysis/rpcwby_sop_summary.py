"""Parse the RPCWBY SOP summary sheets - measured SOP at two temperatures over aging.

WHY THIS SHEET MATTERS MORE THAN THE RAW CSVs
    Test#1 and Test#2 are 8-day, 700k-row aging records in which the chamber
    alternates between 10 and 25 degC. The summary workbook already reduces them
    to what is wanted here: measured SOP per SOC, per temperature, per cycle,
    with the cell's capacity recorded alongside. Re-deriving that from the raw
    files would add nothing but a chance to get it wrong.

WHAT IT ANSWERS
    Whether theta(SOC, SOH, T) SEPARATES - i.e. whether the ratio between the
    10 degC and 25 degC response stays fixed as the cell ages. The ECM+KF plan
    (docs/ecm_kf_plan.md section 3.5.2) has to assume something about this,
    because UYPYDJ ages at a single temperature. Test#1/#2 let it be MEASURED at
    two temperatures instead of assumed - which is the whole reason for having
    more than one dataset.

SHEET LAYOUT (read from the file, not assumed)
    Blocks repeat down the sheet. Each begins with "Cycle N" in column B and
    contains:
        row +2   "25degC" in col B, "10degC ..." in col F - the second may carry
                 the measured capacity as trailing text, e.g. "10°C 2.911126"
        row +3   SOP_disch | SOP_disch_CCCV | SOP_char | SOP_char_CCCV, twice
        rows     SOC in col A, then the eight value columns
    Columns from I onward hold a separate HPPC table and are not read here.

SIGN
    SOP_disch is reported negative (discharge). It is kept signed, and any
    magnitude comparison takes the absolute value explicitly.
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "raw", "RPCWBY", "0_SOP_summary.xlsx")
OUT = os.path.join(HERE, "rpcwby_sop_summary.csv")

FIELDS = ("SOP_disch", "SOP_disch_CCCV", "SOP_char", "SOP_char_CCCV")


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def parse_sheet(ws, sheet):
    rows = list(ws.iter_rows(values_only=True))
    out = []
    starts = [i for i, r in enumerate(rows)
              if len(r) > 1 and isinstance(r[1], str)
              and re.fullmatch(r"\s*Cycle\s*\d+\s*", r[1] or "")]
    for bi, i in enumerate(starts):
        cyc = int(re.search(r"\d+", rows[i][1]).group())
        stop = starts[bi + 1] if bi + 1 < len(starts) else len(rows)

        # Temperature labels and any capacity glued to them.
        temps, caps = {}, {}
        for j in range(i, min(i + 5, stop)):
            for col, v in enumerate(rows[j][:9]):
                if not isinstance(v, str):
                    continue
                m = re.match(r"\s*(-?\d+)\s*°?C", v)
                if m:
                    temps[col] = int(m.group(1))
                    mc = re.search(r"(\d\.\d{3,})", v)
                    if mc:
                        caps[col] = float(mc.group(1))
        if len(temps) < 2:
            continue
        cols = sorted(temps)

        for j in range(i, stop):
            soc = _num(rows[j][0]) if rows[j] else None
            if soc is None or not (0.0 <= soc <= 1.0):
                continue
            for c in cols:
                rec = {"sheet": sheet, "cycle": cyc, "temp_C": temps[c],
                       "SOC": round(soc, 4),
                       "cap_Ah": caps.get(c, "")}
                any_v = False
                for k, name in enumerate(FIELDS):
                    v = _num(rows[j][c + k]) if c + k < len(rows[j]) else None
                    rec[name] = "" if v is None else round(v, 4)
                    any_v = any_v or v is not None
                if any_v:
                    out.append(rec)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--out", default=OUT)
    ap.add_argument("--sheets", default="Test#1,Test#2")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    rows = []
    for sn in args.sheets.split(","):
        if sn not in wb.sheetnames:
            print(f"  {sn}: 시트 없음"); continue
        got = parse_sheet(wb[sn], sn)
        cyc = sorted({r["cycle"] for r in got})
        temps = sorted({r["temp_C"] for r in got})
        caps = sorted({r["cap_Ah"] for r in got if r["cap_Ah"] != ""})
        print(f"  {sn}: {len(got)}행, cycle {cyc[:6]}{'...' if len(cyc) > 6 else ''} "
              f"({len(cyc)}개), 온도 {temps}")
        if caps:
            print(f"          용량 {min(caps):.4f}~{max(caps):.4f} Ah "
                  f"-> SOH {min(caps)/max(caps):.3f}~1.000")
        rows += got

    if not rows:
        sys.exit("파싱된 행이 없습니다")
    with open(args.out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"\n{args.out}: {len(rows)}행")


if __name__ == "__main__":
    main()
