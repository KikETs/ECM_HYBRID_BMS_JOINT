"""Held-out check: does a fresh-cell temperature factor still hold on an aged cell?

THE QUESTION
    ecm_surface.py builds theta(SOC, SOH, T) by multiplying a 25 degC aging
    surface (UYPYDJ) by a temperature factor measured on a FRESH cell
    (Mendeley). Nothing in either source says the factor survives aging.

    RPCWBY Test#1 and Test#2 do: they measure SOP at 10 AND 25 degC at the same
    cycle, from fresh down to about 80 % SOH, on cells neither of the other two
    datasets contains.

THE TEST, WHICH DELIBERATELY AVOIDS CELL-TO-CELL DIFFERENCES
    Absolute resistance differs between cells - at SOH 0.75 the six UYPYDJ cells
    span 5.58x in R1 alone. So the check does NOT compare absolute values across
    datasets. It takes Test#1's OWN measured R at 25 degC, applies the Mendeley
    factor, and compares against Test#1's OWN measured R at 10 degC:

        predicted R10 = measured R25 * g(SOC, |I|, 10 degC)

    Any error is then the factor's, not the cell's.

RESISTANCE IS BACK-CALCULATED FROM THE MEASURED SOP
    R = (OCV - V) / |I| with V = SOP / I. This reproduces the sheet's own
    R_dish column to within 0.2 mOhm wherever both are at the same current, so
    the inversion is sound.

    ONLY POINTS WHERE BOTH TEMPERATURES ARE CURRENT-LIMITED AT 30 A ARE USED.
    Once the cold cell drops to voltage limiting its SOP is set at a different
    current, and comparing the two would confound the temperature effect with a
    current effect - the mistake that first made this look like a much larger
    disagreement than it is.
"""
from __future__ import annotations

import argparse
import os
import re
import sys

import numpy as np
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ecm_surface import TempFactor  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "raw", "RPCWBY", "0_SOP_summary.xlsx")

I_LIM = 30.0
V_MIN = 2.55
T_COLD, T_WARM = 10.0, 25.0


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def read_blocks(ws):
    """(cycle, SOH, SOC, R25, R10) for usable points."""
    rows = list(ws.iter_rows(values_only=True))
    cap = {}
    for r in rows:
        a = _num(r[21]) if len(r) > 21 else None
        b = _num(r[22]) if len(r) > 22 else None
        if a and b:
            cap[int(a)] = b
    cap0 = max(cap.values()) if cap else None

    starts = [i for i, r in enumerate(rows)
              if len(r) > 1 and isinstance(r[1], str)
              and re.fullmatch(r"\s*Cycle\s*\d+\s*", r[1] or "")]
    out = []
    for bi, i in enumerate(starts):
        cyc = int(re.search(r"\d+", rows[i][1]).group())
        stop = starts[bi + 1] if bi + 1 < len(starts) else len(rows)
        soh = cap[cyc] / cap0 if (cap0 and cyc in cap) else np.nan
        # The HPPC block is NOT row-aligned with the SOP block in every cycle,
        # so OCV is looked up by its own SOC rather than by row position.
        ocv = {}
        for j in range(i, stop):
            r = rows[j]
            s = _num(r[9]) if len(r) > 9 else None
            o = _num(r[10]) if len(r) > 10 else None
            if s is not None and o is not None and 0 < s <= 1:
                ocv[round(s, 4)] = o
        for j in range(i, stop):
            r = rows[j]
            soc = _num(r[0])
            if soc is None or not 0 < soc <= 1:
                continue
            o = ocv.get(round(soc, 4))
            s25, s10 = _num(r[1]), _num(r[5])
            if None in (o, s25, s10):
                continue
            v25, v10 = s25 / -I_LIM, s10 / -I_LIM
            if v25 < V_MIN or v10 < V_MIN:        # not both current-limited
                continue
            r25, r10 = (o - v25) / I_LIM, (o - v10) / I_LIM
            if r25 <= 0 or r10 <= 0:
                continue
            out.append((cyc, soh, soc, r25 * 1000, r10 * 1000))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--sheets", default="Test#1,Test#2")
    args = ap.parse_args()

    g = TempFactor()
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    for sn in args.sheets.split(","):
        if sn not in wb.sheetnames:
            continue
        pts = read_blocks(wb[sn])
        if not pts:
            print(f"{sn}: 사용 가능한 점 없음"); continue
        a = np.array(pts)
        gv = np.array([g(p[2], I_LIM, T_COLD)[0] for p in pts])
        pred = a[:, 3] * gv
        err = (pred - a[:, 4]) / a[:, 4] * 100

        print(f"=== {sn}: {len(a)}점 (양쪽 30 A 전류제한) ===")
        print(f"  예측 g(10 °C, 30 A) = {gv.mean():.3f} ± {gv.std():.3f}")
        print(f"  전체 오차: 중앙 {np.median(err):+.1f} %, "
              f"평균절대 {np.abs(err).mean():.1f} %, 95 % {np.percentile(np.abs(err),95):.1f} %")
        print(f"\n  {'SOH 구간':>12} {'n':>4} {'실측 R10/R25':>13} {'예측 g':>8} {'오차%':>8}")
        for lo, hi in ((0.97, 1.01), (0.94, 0.97), (0.91, 0.94),
                       (0.88, 0.91), (0.85, 0.88), (0.70, 0.85)):
            m = (a[:, 1] >= lo) & (a[:, 1] < hi)
            if m.sum() < 3:
                continue
            print(f"  {lo:.2f}~{hi:.2f}   {m.sum():>4} "
                  f"{np.median(a[m, 4] / a[m, 3]):>13.3f} {gv[m].mean():>8.3f} "
                  f"{np.median(err[m]):>+7.1f}")
        print()


if __name__ == "__main__":
    main()
