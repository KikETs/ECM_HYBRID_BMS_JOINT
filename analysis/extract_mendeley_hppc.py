"""Extract the Mendeley 30T HPPC sheet into a tidy (SOC x T x C-rate) resistance grid.

WHY A SCRIPT AND NOT A ONE-OFF
    The sheet is a wide layout with one 18-column block per temperature, and the
    pulse C-rates DIFFER BETWEEN TEMPERATURES (0.5/1/2/4 at -20 degC against
    1/2/6/12 at 25 degC - the protocol reduces current when cold). Reading it by
    hand invites silently pairing a -20 degC 0.5C column with a 25 degC 1C one.
    The block layout is parsed from the header rows instead of hardcoded.

LAYOUT (per temperature block, offsets from the block's OCV column)
    +0        OCV (V)
    +1..+4    Discharge pulse VOLTAGE at four C-rates
    +5..+8    Charge pulse VOLTAGE at four C-rates
    +9..+12   Discharge pulse RESISTANCE at the same four C-rates
    +13..+16  Charge pulse RESISTANCE at the same four C-rates
    Row 71 carries the C-rate for every one of those columns, so the rate is
    read per column rather than assumed.

OUTPUT
    mendeley_hppc_resistance.csv  - one row per (T, SOC, direction, C-rate)
    Resistance is reported in mOhm exactly as the sheet gives it; no rescaling.
"""
import csv
import sys

import openpyxl

SRC = "raw/Mendeley/IN21700-30T HPPC OCV CC_DIS Data.xlsx"
OUT = "analysis/mendeley_hppc_resistance.csv"
SHEET = "HPPC Resistance"
ROW_TEMP, ROW_RATE, ROW_KIND, ROW_DATA0 = 69, 71, 72, 73


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]

    # Temperature block starts, from the header row itself.
    blocks = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(ROW_TEMP, c).value
        if isinstance(v, str) and "C" in v:
            blocks.append((c, float(v.replace("⁰C", "").replace("C", "").strip())))
    print(f"temperature blocks: {[(c, t) for c, t in blocks]}")

    # The 'kind' row labels which run of columns is voltage and which is resistance.
    kinds = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(ROW_KIND, c).value
        if isinstance(v, str) and "Pulse Res" in v:
            kinds[c] = "charge" if v.startswith("Charge") else "discharge"

    rows = []
    for bcol, temp in blocks:
        # resistance runs belonging to this block: the two nearest kind-labels at
        # or after the OCV column, before the next block starts
        nxt = min([c for c, _ in blocks if c > bcol], default=ws.max_column + 1)
        runs = [(c, k) for c, k in kinds.items() if bcol <= c < nxt]
        for start, direction in sorted(runs):
            for off in range(4):
                col = start + off
                rate = ws.cell(ROW_RATE, col).value
                if not isinstance(rate, (int, float)):
                    continue
                for r in range(ROW_DATA0, ws.max_row + 1):
                    soc = ws.cell(r, 1).value
                    ah = ws.cell(r, 2).value
                    ocv = ws.cell(r, bcol).value
                    res = ws.cell(r, col).value
                    if not isinstance(soc, (int, float)) or not isinstance(res, (int, float)):
                        continue
                    rows.append({
                        "T_C": temp,
                        "SOC": soc / 100.0,
                        "Ah": ah if isinstance(ah, (int, float)) else "",
                        "Direction": direction,
                        "CRate": float(rate),
                        "OCV_V": ocv if isinstance(ocv, (int, float)) else "",
                        "R_mOhm": res,
                    })

    if not rows:
        sys.exit("no rows extracted - the sheet layout changed")
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {OUT}: {len(rows)} rows")

    temps = sorted({r["T_C"] for r in rows})
    print("\ncoverage")
    for t in temps:
        sub = [r for r in rows if r["T_C"] == t]
        for d in ("discharge", "charge"):
            s = [r for r in sub if r["Direction"] == d]
            if not s:
                continue
            rates = sorted({r["CRate"] for r in s})
            socs = sorted({r["SOC"] for r in s})
            print(f"  T={t:>5} {d:<9} C-rates {rates}  SOC {min(socs):.3f}-{max(socs):.3f} ({len(socs)}pt)")


if __name__ == "__main__":
    main()
