"""Test#3 시트: 저자들이 뽑은 SOP 를 세 지평 x 여섯 온도로.

WHY THIS SHEET AND NOT THE RAW PULSES
    18.3 절은 Test#3 의 원시 펄스에서 SOP 라벨을 만들 수 없다고 결론지었다 —
    탐색이 수렴하므로 (T, tau, SOC) 마다 전류 하나만 남고 팬이 없다. 그런데
    저자들은 그 수렴값 자체를 SOP 로 기록해 두었고, 그것이 이 시트다. 탐색의
    수렴값은 외삽이 아니라 측정이다.

    rpcwby_sop_summary.py 는 Test#1/#2 만 돌았다(--sheets 기본값). Test#3 은
    레이아웃이 달라 그 파서로는 읽히지 않는다.

레이아웃 (시트에서 읽은 것, 가정하지 않음)
    행 0   "SOP Limits: | Current (+-0.1A) | n15 to 30 | Voltage (+-0.04V) | 2.55 to 4.15"
    행 1   열 A "2-second", 열 F "10-second", 열 K "30-second"  <- 세 개의 열 밴드
    각 밴드 안에서 블록이 아래로 반복한다:
        헤더행   c+0 "SOC",  c+1 "40°C - 2.6333Ah"  (용량이 붙는 경우가 있다)
        이름행   c+1..c+4  SOP_disch | SOP_disch_CCCV | SOP_char | SOP_char_CCCV
        자료행   c+0 = SOC,  c+1..c+4 = 값
    블록은 다음 "SOC" 헤더까지다.

왜 tau 가 중요한가
    19.3 절에서 tau = 2 s 는 필요한 저항 배수가 1.323(90 %tile 1.843)로 tau=10 s
    의 1.076 과 전혀 다르고, 21 절의 무초과 여유가 tau=2 s 에서 0.489 — 즉 절반을
    깎아야 한다. 그런데 이 프로젝트의 tau=2 s 행은 **전부 외삽**이었다
    (extrap<=1.0 인 143 행이 전부 tau=10 s). 이 시트가 처음으로 측정된 tau=2 s
    SOP 를 준다.
"""
from __future__ import annotations

import argparse
import csv
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "raw", "RPCWBY", "0_SOP_summary.xlsx")
OUT = os.path.join(HERE, "rpcwby_sop_test3.csv")
FIELDS = ("SOP_disch", "SOP_disch_CCCV", "SOP_char", "SOP_char_CCCV")


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


def band_columns(rows):
    """행 1 의 '<n>-second' 라벨에서 밴드 시작 열과 지평을 읽는다."""
    out = {}
    for j in range(min(4, len(rows))):
        for col, v in enumerate(rows[j]):
            if isinstance(v, str):
                m = re.fullmatch(r"\s*(\d+)\s*-\s*second\s*", v, re.I)
                if m:
                    out[col] = float(m.group(1))
    return dict(sorted(out.items()))


def parse_bands(ws, sheet):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    bands = band_columns(rows)
    if not bands:
        return [], {}
    out = []
    for c0, tau in bands.items():
        # 이 밴드에서 "SOC" 헤더가 있는 행들
        heads = [j for j, r in enumerate(rows)
                 if c0 < len(r) and isinstance(r[c0], str)
                 and r[c0].strip().upper() == "SOC"]
        for hi, j0 in enumerate(heads):
            lbl = rows[j0][c0 + 1] if c0 + 1 < len(rows[j0]) else None
            if not isinstance(lbl, str):
                continue
            m = re.search(r"(-?\d+)\s*°?\s*C", lbl)
            if not m:
                continue
            temp = int(m.group(1))
            mc = re.search(r"(\d\.\d{3,})", lbl)
            cap = float(mc.group(1)) if mc else ""
            stop = heads[hi + 1] if hi + 1 < len(heads) else len(rows)
            for j in range(j0 + 1, stop):
                soc = _num(rows[j][c0]) if c0 < len(rows[j]) else None
                if soc is None or not (0.0 <= soc <= 1.0):
                    continue
                rec = {"sheet": sheet, "tau_s": tau, "temp_C": temp,
                       "SOC": round(soc, 4), "cap_Ah": cap}
                any_v = False
                for k, name in enumerate(FIELDS):
                    v = _num(rows[j][c0 + 1 + k]) if c0 + 1 + k < len(rows[j]) else None
                    rec[name] = "" if v is None else round(v, 4)
                    any_v = any_v or v is not None
                if any_v:
                    out.append(rec)
    return out, bands


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--out", default=OUT)
    ap.add_argument("--sheets", default="Test#3")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.xlsx, data_only=True, read_only=True)
    rows = []
    for sn in a.sheets.split(","):
        if sn not in wb.sheetnames:
            print(f"  {sn}: 시트 없음"); continue
        got, bands = parse_bands(wb[sn], sn)
        rows += got
        print(f"  {sn}: 밴드 {[f'{chr(65+c)}={t:.0f}s' for c, t in bands.items()]}")
        for tau in sorted({r['tau_s'] for r in got}):
            g = [r for r in got if r["tau_s"] == tau]
            ts = sorted({r["temp_C"] for r in g})
            nd = sum(1 for r in g if r["SOP_disch"] != "")
            print(f"    tau={tau:>4.0f}s  {len(g):>3}행  방전값 {nd:>3}  온도 {ts}")
        caps = sorted({r["cap_Ah"] for r in got if r["cap_Ah"] != ""})
        if caps:
            print(f"    기록된 용량 {caps}  -> SOH {min(caps)/3.0:.3f}~{max(caps)/3.0:.3f} (정격 3.0 Ah)")
    if not rows:
        return
    with open(a.out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"\n  {a.out}  {len(rows)}행")


if __name__ == "__main__":
    main()
